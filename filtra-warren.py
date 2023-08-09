import sys
import webbrowser
from pathlib import Path
from pprint import pprint

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


def abrir_arquivo(caminho):
    url = f'file://{caminho}'
    webbrowser.open(url)


def printd(args):
    return print(f'{args=}')


ROOT_FOLDER = Path().cwd()

printd(ROOT_FOLDER)

WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

printd(WORKBOOK_PATH)

# nome para a planilha
SHEET_NAME = 'Minha planilha'

# Criamos a planilha
# workbook.create_sheet(sheet_name)

# Carregando um arquivo do excel
try:
    workbook: Workbook = load_workbook(WORKBOOK_PATH)
except Exception as e:
    workbook: Workbook = Workbook()
    workbook.create_sheet(SHEET_NAME)

# Seleciou a planilha
worksheet: Worksheet = workbook[SHEET_NAME]

FILTRO_NOME_ATIVOS = ['Fundo', 'Warren', 'CDB', 'LC', 'CRI', 'CRA', 'MS', 'Deb', 'Tesouro']
FILTRO_CONTINUE = ['RENDA FIXA', 'RENDA VARIÁVEL', 'OUTROS', 'Percentual de alocação']

lista_ = list()
lista_.append(['Fundos', 'Valor R$'])

soma_total = 0
nome_ativo = None
pegar_proximo_valor = False
pular = False

row: tuple[Cell]
for row in worksheet.iter_rows():
    for cell in row:
        texto_celula = str(cell.value)

        if 'Saldo bruto' in texto_celula:
            pegar_proximo_valor = True
            continue

        if pegar_proximo_valor:
            pegar_proximo_valor = False
            valor_ativo = str(cell.value)

            if "R$" in str(cell.value):
                valor_ativo = valor_ativo[3:]  # Removendo R$ e NBSP
                # R  $  NBSP
                # [0, 1, 2]
                valor_ativo = valor_ativo.replace('.', '').replace(',', '.')

            valor_ativo = float(valor_ativo)
            soma_total += valor_ativo
            lista_.append([nome_ativo, valor_ativo])
            continue

        try:  # talvez de um bug com ativos que tenham ativos com menos de 1 real
            if float(texto_celula) < 1:
                continue
        except ValueError:
            pass

        for item in FILTRO_CONTINUE:
            if item in texto_celula:
                pular = True
                break

        if pular:
            pular = False
            continue

        for correspondencia in FILTRO_NOME_ATIVOS:
            if correspondencia in texto_celula:
                if correspondencia == 'Fundo':
                    nome_ativo = texto_celula.replace('Fundo ', '')
                else:
                    nome_ativo = texto_celula
                break

lista_.append(['Total', soma_total])

printd(lista_)

# criando uma nova aba
NOVA_ABA = 'Filtrado'
workbook.create_sheet(NOVA_ABA)
worksheet = workbook[NOVA_ABA]

for ativo in lista_:
    worksheet.append(ativo)

not_exclude = ['Minha planilha', 'Filtrado']

for tabela in workbook.sheetnames:
    if tabela not in not_exclude:
        workbook.remove(workbook[tabela])

workbook.save(WORKBOOK_PATH)
abrir_arquivo(WORKBOOK_PATH)
input('Pressione qualquer tecla para sair + Enter.')
