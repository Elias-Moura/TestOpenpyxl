from pathlib import Path
from pprint import pprint

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent

WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

# nome para a planilha
SHEET_NAME = 'Minha planilha'

# Criamos a planilha
# workbook.create_sheet(sheet_name)

# Carregando um arquivo do excel
try:
    workbook: Workbook = load_workbook(WORKBOOK_PATH)
except Exception as e:
    workbook: Workbook = Workbook()
    workbook.create_sheet(SHEET_NAME)

# Seleciou a planilha
worksheet: Worksheet = workbook[SHEET_NAME]

FILTRO_NOME_ATIVOS = ['Fundo', 'Warren', 'CDB', 'LC', 'CRI', 'CRA', 'MS', 'Deb', 'Tesouro']

lista_ = list()
lista_.append(['Fundos', 'Valor R$'])

soma_total = 0
nome_ativo = None

row: tuple[Cell]
for row in worksheet.iter_rows():
    for cell in row:
        # print(cell.value)
        for correspondencia in FILTRO_NOME_ATIVOS:
            texto = str(cell.value)
            if correspondencia in texto:
                if correspondencia == 'Fundo' or 'Fundo Warren Tesouro SELIC':
                    nome_ativo = texto.replace('Fundo ', '')
                else:
                    nome_ativo = texto
                break

        if "R$" in str(cell.value):
            valor_ativo = str(cell.value)[3:]  # Removendo R$ e NBSP
            # R  $  NBSP
            # [0, 1, 2]
            valor_ativo = valor_ativo.replace('.', '').replace(',', '.')
            valor_ativo = float(valor_ativo)
            soma_total += valor_ativo
            lista_.append([nome_ativo, valor_ativo])

lista_.append(['Total', soma_total])

pprint(lista_)

# criando uma nova aba
NOVA_ABA = 'Filtrado'
workbook.create_sheet(NOVA_ABA)
worksheet = workbook[NOVA_ABA]

for ativo in lista_:
    worksheet.append(ativo)

# not_exclude = ['Minha planilha', 'Filtrado']
#
# for tabela in workbook.sheetnames:
#     if tabela not in not_exclude:
#         workbook.remove(workbook[tabela])

workbook.save(WORKBOOK_PATH)
